#!/usr/bin/env python3
"""
NSC Voting Results Converter

Converts host-provided Excel result files (varying formats) into the
standardized unpivoted format: Edition, Subevent, Voter, Nation, Points

Usage:
    python nsc_votes_converter.py NSC247.xlsx --edition 247
    python nsc_votes_converter.py NSC247.xlsx --edition 247 --append votes_master.xlsx
    python nsc_votes_converter.py NSC247.xlsx --edition 247 --output votes_247.xlsx
"""

import argparse
import json
import logging
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("votes_converter")

# ── Sheet name → subevent classification ───────────────────────

SHEET_PATTERNS = [
    # (pattern, subevent) — checked in order, first match wins
    # SF1 Regular
    (r"s(?:emi[- ]?(?:final)?)?[- ]?1[- ]?(?:reg|regular|voting)?$", "S1"),
    (r"sf[- ]?1[- ]?(?:reg|regular|voting)?$", "S1"),
    (r"s1\s*reg", "S1"),
    # SF1 REJU
    (r"s(?:emi[- ]?(?:final)?)?[- ]?1[- ]?(?:rej|reju|rj)", "R1"),
    (r"sf[- ]?1[- ]?(?:rej|reju|rj)", "R1"),
    (r"reju[- ]?1", "R1"),
    (r"s1\s*rej", "R1"),
    # SF2 Regular
    (r"s(?:emi[- ]?(?:final)?)?[- ]?2[- ]?(?:reg|regular|voting)?$", "S2"),
    (r"sf[- ]?2[- ]?(?:reg|regular|voting)?$", "S2"),
    (r"s2\s*reg", "S2"),
    # SF2 REJU
    (r"s(?:emi[- ]?(?:final)?)?[- ]?2[- ]?(?:rej|reju|rj)", "R2"),
    (r"sf[- ]?2[- ]?(?:rej|reju|rj)", "R2"),
    (r"reju[- ]?2", "R2"),
    (r"s2\s*rej", "R2"),
    # Grand Final
    (r"^final$", "GF"),
    (r"^grand\s*final$", "GF"),
    (r"^final\s*(?:results|voting|full)?$", "GF"),
    # Waiting List
    (r"wl", "WL"),
    (r"waiting\s*list", "WL"),
    (r"waliju", "WL"),
    # Semi results (combined) — skip
    (r"semi\s*results", "_SKIP"),
    (r"nq\s*results", "_SKIP"),
    # Non-voting sheets
    (r"entries", "_SKIP"),
    (r"draw", "_SKIP"),
    (r"pots", "_SKIP"),
    (r"all\s*entries", "_SKIP"),
    (r"sf[12]\s*entries", "_SKIP"),
]

# Columns that are metadata, NOT voter names
META_COLS = {
    "draw", "place", "country", "nation", "pot", "total", "points",
    "artist", "song", "rank", "placing", "running order", "number",
    "link", "video", "video link", "recap", "recap timestamps",
    "semi", "user", "", "bonus",
}

# Also skip columns that look like formulas or computed values
FORMULA_PATTERNS = [r"^=", r"^total$", r"^points$", r"^placing$", r"^rank$"]


def classify_sheet(name: str) -> str | None:
    """Classify a sheet name into a subevent code or None to skip."""
    clean = name.lower().strip()
    for pattern, subevent in SHEET_PATTERNS:
        if re.search(pattern, clean):
            return subevent if subevent != "_SKIP" else None
    return None


def is_meta_column(col_name: str) -> bool:
    """Check if a column header is metadata (not a voter name)."""
    if not col_name:
        return True
    clean = col_name.lower().strip()
    if clean in META_COLS:
        return True
    if any(re.match(p, clean) for p in FORMULA_PATTERNS):
        return True
    # Pure numbers are draw/place columns
    if re.match(r"^\d+\.?\d*$", clean):
        return True
    # Single character or very short strings that look like abbreviations for stats
    if len(clean) <= 1 and clean not in ("ÿ",):  # "Ÿ" is a nation name
        return True
    return False


def safe_str(v) -> str:
    """Safely convert a value to a trimmed string."""
    if v is None:
        return ""
    s = str(v).strip()
    # Remove leading/trailing quotes
    s = s.strip("'\"")
    return s


def safe_int(v) -> int | None:
    """Try to extract an integer from a cell value."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v) if isinstance(v, float) and v == int(v) else v
        return int(n) if isinstance(n, (int, float)) and n > 0 else None
    s = str(v).strip()
    s = re.sub(r"[^\d]", "", s)
    return int(s) if s else None


def find_header_row(ws, max_scan=10) -> tuple[int, list[str]]:
    """
    Find the row that contains the voter column headers.
    Returns (row_index, list_of_header_values).
    
    Heuristic: the header row has many non-empty, non-numeric cells
    that look like nation names (strings with length > 2).
    """
    best_row = 0
    best_score = 0
    best_headers = []
    
    rows_data = []
    for i, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        rows_data.append(list(row))
    
    for i, row in enumerate(rows_data):
        score = 0
        for cell in row:
            s = safe_str(cell)
            if len(s) > 2 and not re.match(r"^\d+\.?\d*$", s) and not s.startswith("="):
                score += 1
        if score > best_score:
            best_score = score
            best_row = i
            best_headers = [safe_str(c) for c in row]
    
    return best_row, best_headers


def parse_voting_sheet(ws, subevent: str) -> list[dict]:
    """
    Parse a voting matrix sheet into unpivoted rows.
    Returns list of {voter, nation, points}.
    """
    # Read all rows
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    
    if len(all_rows) < 2:
        return []
    
    # Find header row
    header_row_idx, headers = find_header_row(ws)
    
    if not headers:
        log.warning(f"    Could not find header row")
        return []
    
    # Identify nation column (first column with nation-like names)
    # And voter columns (non-metadata columns after nation)
    nation_col = None
    voter_cols = []  # (col_index, voter_name)
    
    for col_idx, h in enumerate(headers):
        h_lower = h.lower().strip()
        
        if nation_col is None:
            # Check if this column header explicitly marks it
            if h_lower in ("country", "nation"):
                nation_col = col_idx
                continue
            # Skip known metadata headers
            if h_lower in ("draw", "place", "pot", "rank", "placing", "running order", "number", "#"):
                continue
            # Empty header or unrecognized: check what data looks like
            if h_lower == "" or is_meta_column(h):
                # Peek at data rows to decide
                sample_vals = []
                for r in range(header_row_idx + 1, min(header_row_idx + 6, len(all_rows))):
                    if r < len(all_rows) and col_idx < len(all_rows[r]):
                        sample_vals.append(safe_str(all_rows[r][col_idx]))
                
                # If most values are numbers, this is a draw/place column — skip
                num_count = sum(1 for v in sample_vals if re.match(r"^\d+\.?\d*$", v))
                name_count = sum(1 for v in sample_vals if len(v) > 2 and not re.match(r"^\d+\.?\d*$", v))
                
                if name_count >= 3:
                    nation_col = col_idx
                    continue
                else:
                    continue  # skip this column, keep looking
            
            # Non-empty header that's not metadata: check if data has nation-like names
            sample_vals = [safe_str(all_rows[r][col_idx]) if r < len(all_rows) and col_idx < len(all_rows[r]) else ""
                          for r in range(header_row_idx + 1, min(header_row_idx + 5, len(all_rows)))]
            name_like = sum(1 for v in sample_vals if len(v) > 2 and not re.match(r"^\d+\.?\d*$", v))
            if name_like >= 2:
                nation_col = col_idx
                continue
        
        # After finding nation column, identify voter columns
        if nation_col is not None and col_idx > nation_col:
            if not is_meta_column(h):
                voter_cols.append((col_idx, h))
    
    if nation_col is None:
        # Fallback: assume first column is nation
        nation_col = 0
        # If first column has numbers, try second
        sample = safe_str(all_rows[header_row_idx + 1][0] if header_row_idx + 1 < len(all_rows) else "")
        if re.match(r"^\d+\.?\d*$", sample):
            nation_col = 1
        
        # Re-scan for voter columns
        voter_cols = []
        for col_idx, h in enumerate(headers):
            if col_idx > nation_col and not is_meta_column(h):
                voter_cols.append((col_idx, h))
    
    if not voter_cols:
        log.warning(f"    No voter columns found. Headers: {headers[:15]}")
        return []
    
    log.info(f"    Nation col: {nation_col}, Voters: {len(voter_cols)}, Header row: {header_row_idx}")
    
    # Parse data rows
    results = []
    for row_idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[row_idx]
        if not row or nation_col >= len(row):
            continue
        
        nation = safe_str(row[nation_col])
        if not nation or len(nation) < 2:
            continue
        # Skip if nation looks like a number
        if re.match(r"^\d+\.?\d*$", nation):
            continue
        
        for col_idx, voter_name in voter_cols:
            if col_idx >= len(row):
                continue
            pts = safe_int(row[col_idx])
            if pts is not None and pts > 0:
                results.append({
                    "voter": voter_name,
                    "nation": nation,
                    "points": pts,
                })
    
    return results


def normalize_name(name: str) -> str:
    """Normalize a nation name: proper case, trim, clean."""
    if not name:
        return name
    # Strip whitespace
    name = name.strip()
    # Some files use ALL CAPS for voter names
    if name.isupper() and len(name) > 3:
        # Convert to title case but preserve known special cases
        name = name.title()
    # Fix common casing issues
    name = re.sub(r"\s+", " ", name)  # collapse spaces
    return name


def process_file(filepath: str, edition: int) -> list[dict]:
    """
    Process a single host Excel file and return all unpivoted vote rows.
    Each row: {edition, subevent, voter, nation, points}
    """
    log.info(f"Processing: {os.path.basename(filepath)} (Edition #{edition})")
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_votes = []
    
    for sheet_name in wb.sheetnames:
        subevent = classify_sheet(sheet_name)
        if subevent is None:
            log.info(f"  Skipping sheet: '{sheet_name}'")
            continue
        
        log.info(f"  Processing sheet: '{sheet_name}' → {subevent}")
        ws = wb[sheet_name]
        
        rows = parse_voting_sheet(ws, subevent)
        
        for r in rows:
            all_votes.append({
                "Edition": edition,
                "Subevent": subevent,
                "Voter": normalize_name(r["voter"]),
                "Nation": normalize_name(r["nation"]),
                "Points": r["points"],
            })
        
        log.info(f"    → {len(rows)} vote records")
    
    wb.close()
    
    # Summary
    subevents = defaultdict(int)
    for v in all_votes:
        subevents[v["Subevent"]] += 1
    
    log.info(f"  Total: {len(all_votes)} votes")
    for sub, count in sorted(subevents.items()):
        log.info(f"    {sub}: {count}")
    
    return all_votes


def load_existing(filepath: str) -> list[dict]:
    """Load existing master votes file."""
    if not os.path.exists(filepath):
        return []
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Sheet1"]
    headers = None
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) for h in row]
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    log.info(f"Loaded {len(rows)} existing records from {filepath}")
    return rows


def save_xlsx(votes: list[dict], filepath: str):
    """Save votes to Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Edition", "Subevent", "Voter", "Nation", "Points"])
    for v in votes:
        ws.append([v["Edition"], v["Subevent"], v["Voter"], v["Nation"], v["Points"]])
    wb.save(filepath)
    log.info(f"Saved {len(votes)} records to {filepath}")


def save_csv(votes: list[dict], filepath: str):
    """Save votes to CSV."""
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("Edition,Subevent,Voter,Nation,Points\n")
        for v in votes:
            voter = v["Voter"].replace('"', '""')
            nation = v["Nation"].replace('"', '""')
            f.write(f'{v["Edition"]},{v["Subevent"]},"{voter}","{nation}",{v["Points"]}\n')
    log.info(f"Saved {len(votes)} records to {filepath}")


def main():
    parser = argparse.ArgumentParser(description="NSC Voting Results Converter")
    parser.add_argument("input", help="Input Excel file from host")
    parser.add_argument("--edition", type=int, required=True, help="Edition number")
    parser.add_argument("--output", default=None, help="Output file path (xlsx or csv)")
    parser.add_argument("--append", default=None, help="Append to existing master file (xlsx)")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()
    
    if args.verbose:
        log.setLevel(logging.DEBUG)
    
    # Process input file
    votes = process_file(args.input, args.edition)
    
    if not votes:
        log.error("No votes extracted! Check the file format.")
        return 1
    
    # Append to master file
    if args.append:
        existing = load_existing(args.append)
        # Remove existing data for this edition (replace)
        existing = [v for v in existing if v.get("Edition") != args.edition]
        combined = existing + votes
        combined.sort(key=lambda v: (v["Edition"], v["Subevent"], v["Voter"], v["Nation"]))
        save_xlsx(combined, args.append)
    
    # Save output
    if args.output:
        ext = os.path.splitext(args.output)[1].lower()
        if ext == ".csv":
            save_csv(votes, args.output)
        else:
            save_xlsx(votes, args.output)
    elif not args.append:
        # Default: save to nsc_{edition}_votes.xlsx
        outpath = f"nsc_{args.edition}_votes.xlsx"
        save_xlsx(votes, outpath)
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"Edition #{args.edition}: {len(votes)} vote records extracted")
    subs = defaultdict(int)
    voters = defaultdict(set)
    for v in votes:
        subs[v["Subevent"]] += 1
        voters[v["Subevent"]].add(v["Voter"])
    for sub in sorted(subs):
        print(f"  {sub}: {subs[sub]} votes from {len(voters[sub])} voters")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
