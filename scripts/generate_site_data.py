#!/usr/bin/env python3
"""
Regenerate all site data files from the scraper output.
Run after nsc_scraper.py to update the website.

Usage:
    python generate_site_data.py --history data/nsc_all_history.json --output ../public/data/
    python generate_site_data.py --history data/nsc_all_history.json --votes NSC_Votes_Unpivot.xlsx --output ../public/data/
"""
import argparse, json, os, sys
from collections import defaultdict, Counter

def sn(v):
    if v is None: return None
    if isinstance(v, float) and v == int(v): return int(v)
    if isinstance(v, (int, float)): return v
    return None

def main():
    parser = argparse.ArgumentParser(description="Generate NSC site data files")
    parser.add_argument("--history", required=True, help="Path to nsc_all_history.json from scraper")
    parser.add_argument("--votes", default=None, help="Path to NSC_Votes_Unpivot.xlsx (optional)")
    parser.add_argument("--roster", default=None, help="Path to NSCRoster.xlsx (optional, for owner mapping)")
    parser.add_argument("--output", default="../public/data/", help="Output directory")
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)

    with open(args.history, encoding="utf-8") as f:
        raw = json.load(f)
    print(f"Loaded {len(raw)} history entries")

    nsc = [r for r in raw if r.get("Event") == "NSC"]
    gf = [r for r in nsc if r["Subevent"] == "GF"]
    sf = [r for r in nsc if r["Subevent"] in ("SF1", "SF2")]
    sub_map = {"GF":0,"SF1":1,"SF2":2,"MICROSTATE QUALIFICATION":3}

    # ── database.json ──
    db = []
    for r in nsc:
        db.append([r.get("Edition"), sub_map.get(r["Subevent"],0), r.get("Draw"),
            r.get("Nation",""), r.get("Artist",""), r.get("Song",""),
            r.get("Place"), r.get("Points"), r.get("YouTube")])
    with open(os.path.join(args.output, "database.json"), "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, separators=(",",":"))
    print(f"  database.json: {len(db)} rows")

    # ── homepage.json ──
    winners = sorted([r for r in gf if r.get("Place")==1], key=lambda x: x["Edition"])
    gf_sorted = sorted(gf, key=lambda x: (x.get("Points") or 0), reverse=True)
    artist_counts = Counter(r["Artist"] for r in gf if r.get("Artist"))
    win_counts = Counter(r["Nation"] for r in winners)
    latest = winners[-1] if winners else {}
    max_ed = max(r["Edition"] for r in nsc)

    nation_gf = defaultdict(list)
    for r in gf: nation_gf[r["Nation"]].append(r)
    top_nations = []
    for n, count in win_counts.most_common(15):
        entries = nation_gf[n]
        pts = [r.get("Points") for r in entries if r.get("Points")]
        places = [r.get("Place") for r in entries if r.get("Place")]
        top_nations.append({"n":n,"w":count,"e":len(entries),
            "t3":sum(1 for p in places if p<=3),"t5":sum(1 for p in places if p<=5),
            "ap":round(sum(pts)/len(pts),1) if pts else 0,
            "apl":round(sum(places)/len(places),1) if places else 0})

    homepage = {
        "s":{"te":max_ed,"tn":len(set(r["Nation"] for r in nsc)),"gf":len(gf),"sf":len(sf),
            "ue":len(set((r["Edition"],r.get("Artist",""),r.get("Song","")) for r in nsc)),
            "ua":len(set(r["Artist"] for r in gf if r.get("Artist"))),"us":len(set(r["Song"] for r in gf if r.get("Song"))),
            "hs":gf_sorted[0].get("Points"),"hsn":gf_sorted[0]["Nation"],"hse":gf_sorted[0]["Edition"],
            "lw":latest.get("Nation"),"le":latest.get("Edition"),"la":latest.get("Artist"),"ls":latest.get("Song"),"lp":latest.get("Points")},
        "tn":top_nations,
        "rw":[{"e":w["Edition"],"n":w["Nation"],"a":w["Artist"],"s":w["Song"],"p":w.get("Points")} for w in winners[-15:][::-1]],
        "ta":[{"a":a,"c":c} for a,c in artist_counts.most_common(12)],
        "rs":[{"e":r["Edition"],"n":r["Nation"],"a":r["Artist"],"s":r["Song"],"p":r.get("Points")} for r in gf_sorted[:10]],
    }
    with open(os.path.join(args.output, "homepage.json"), "w", encoding="utf-8") as f:
        json.dump(homepage, f, ensure_ascii=False, separators=(",",":"))
    print(f"  homepage.json")

    # ── editions.json ──
    gf_nations_by_ed = defaultdict(set)
    for r in gf: gf_nations_by_ed[r["Edition"]].add(r["Nation"])
    wc = Counter()
    nth_win = {}
    for w in winners: wc[w["Nation"]] += 1; nth_win[(w["Nation"],w["Edition"])] = wc[w["Nation"]]
    total_wins = Counter(w["Nation"] for w in winners)

    editions_d = defaultdict(lambda:{"gf":[],"sf1":[],"sf2":[],"mpq":[]})
    for r in nsc:
        key = {"GF":"gf","SF1":"sf1","SF2":"sf2","MICROSTATE QUALIFICATION":"mpq"}.get(r["Subevent"],"gf")
        editions_d[r["Edition"]][key].append([r.get("Draw"),r["Nation"],r.get("Artist",""),r.get("Song",""),r.get("Place"),r.get("Points"),r.get("YouTube")])

    ed_output = {}
    for ed in sorted(editions_d):
        subs = editions_d[ed]
        gf_s = sorted(subs["gf"], key=lambda x: x[4] if x[4] else 999)
        w = gf_s[0] if gf_s and gf_s[0][4]==1 else None
        ru = gf_s[1] if len(gf_s)>1 else None
        wn = w[1] if w else None
        margin = (w[5]-ru[5]) if w and ru and w[5] and ru[5] else None
        gf_n = set(e[1] for e in subs["gf"])
        r1=r2=None
        for sk,rv in [("sf1","r1"),("sf2","r2")]:
            for e in sorted(subs[sk], key=lambda x: x[4] if x[4] else 999):
                if e[4] and e[4]>10 and e[1] in gf_n:
                    if rv=="r1": r1=e[1]
                    else: r2=e[1]
                    break
        gf_pts = [e[5] for e in subs["gf"] if e[5]]
        ed_output[str(ed)] = {"m":{"w":wn,"wa":w[2] if w else None,"ws":w[3] if w else None,"wp":w[5] if w else None,
            "gs":len(subs["gf"]),"s1":len(subs["sf1"]),"s2":len(subs["sf2"]),"ms":len(subs["mpq"]),
            "mg":margin,"nw":nth_win.get((wn,ed)),"tw":total_wins.get(wn,0),
            "ru":ru[1] if ru else None,"rup":ru[5] if ru else None,
            "ap":round(sum(gf_pts)/len(gf_pts),1) if gf_pts else None,"r1":r1,"r2":r2},
            "gf":subs["gf"],"sf1":subs["sf1"],"sf2":subs["sf2"],"mpq":subs["mpq"]}

    with open(os.path.join(args.output, "editions.json"), "w", encoding="utf-8") as f:
        json.dump(ed_output, f, ensure_ascii=False, separators=(",",":"))
    print(f"  editions.json: {len(ed_output)} editions")

    # ── nations.json ──
    max_gf_place = {}
    for r in gf:
        ed=r["Edition"]; pl=r.get("Place")
        if ed and pl: max_gf_place[ed]=max(max_gf_place.get(ed,0),pl)

    max_sf_place = {}
    for r in sf:
        ed=r["Edition"]; sub=r["Subevent"]; pl=r.get("Place")
        if ed and pl:
            k=(ed,sub)
            max_sf_place[k]=max(max_sf_place.get(k,0),pl)

    # Build REJU qualifier lookup from edition data
    reju_quals = defaultdict(int)  # nation -> count of REJU qualifications
    for ed in sorted(editions_d):
        subs = editions_d[ed]
        gf_n = set(e[1] for e in subs["gf"])
        for sk in ["sf1", "sf2"]:
            for e in sorted(subs[sk], key=lambda x: x[4] if x[4] else 999):
                if e[4] and e[4] > 10 and e[1] in gf_n:
                    reju_quals[e[1]] += 1
                    break

    all_nations = sorted(set(r["Nation"] for r in nsc if r.get("Nation")))
    profiles = {}
    for nation in all_nations:
        n_gf=[r for r in gf if r["Nation"]==nation]
        n_sf=[r for r in sf if r["Nation"]==nation]
        n_all=[r for r in nsc if r["Nation"]==nation]
        if not n_all: continue
        all_eds=sorted(set(r["Edition"] for r in n_all))
        gf_places=[r.get("Place") for r in n_gf if r.get("Place")]
        gf_points=[r.get("Points") for r in n_gf if r.get("Points")]
        wins_n=sum(1 for p in gf_places if p==1)
        sf_q=sum(1 for r in n_sf if r["Nation"] in gf_nations_by_ed.get(r["Edition"],set()))
        gf_last=sum(1 for r in n_gf if r.get("Place") and r.get("Place")==max_gf_place.get(r["Edition"],0))
        sf_last=sum(1 for r in n_sf if r.get("Place") and r.get("Place")==max_sf_place.get((r["Edition"],r["Subevent"]),0))
        gf_eds=sorted(set(r["Edition"] for r in n_gf))
        bs=cs=0
        for ie in range(len(gf_eds)):
            cs=cs+1 if ie==0 or gf_eds[ie]==gf_eds[ie-1]+1 else 1
            bs=max(bs,cs)
        # NQ streak (best and current)
        nq_best=nq_cur=nq_now=0
        ed_map_tmp=defaultdict(dict)
        for r in n_all: ed_map_tmp[r["Edition"]][r["Subevent"]]=r
        for ed_i in all_eds:
            s=ed_map_tmp[ed_i]; sf_r=s.get("SF1") or s.get("SF2"); gf_r=s.get("GF")
            dnq = sf_r is not None and gf_r is None
            if dnq: nq_cur+=1; nq_best=max(nq_best,nq_cur)
            else: nq_cur=0
        # Current NQ streak from end
        for ed_i in reversed(all_eds):
            s=ed_map_tmp[ed_i]; sf_r=s.get("SF1") or s.get("SF2"); gf_r=s.get("GF")
            if sf_r is not None and gf_r is None: nq_now+=1
            else: break
        ac=Counter()
        seen_ed_art=set()
        for r in n_all:
            a=r.get("Artist","")
            if not a: continue
            ek=(r["Edition"],a)
            if ek not in seen_ed_art:
                seen_ed_art.add(ek)
                ac[a]+=1
        ta=ac.most_common(1)[0] if ac else (None,0)
        we=[r["Edition"] for r in n_gf if r.get("Place")==1]
        hist=[]
        for ed in all_eds:
            s=ed_map_tmp[ed]; sf_r=s.get("SF1") or s.get("SF2"); gf_r=s.get("GF"); mpq=s.get("MICROSTATE QUALIFICATION")
            src=gf_r or sf_r or mpq
            hist.append([ed,sf_r["Subevent"] if sf_r else None,sf_r.get("Place") if sf_r else None,sf_r.get("Points") if sf_r else None,
                gf_r.get("Place") if gf_r else None,gf_r.get("Points") if gf_r else None,
                src.get("Artist","") if src else None,src.get("Song","") if src else None])
        profiles[nation]={"n":nation,"gf":len(n_gf),"sf":len(n_sf),"sfQ":sf_q,"sfD":len(n_sf)-sf_q,
            "qr":round(sf_q/len(n_sf)*100,1) if n_sf else None,"te":len(all_eds),"fe":all_eds[0],"le":all_eds[-1],
            "w":wins_n,"we":we,"t3":sum(1 for p in gf_places if p<=3),"t6":sum(1 for p in gf_places if p<=6),
            "t10":sum(1 for p in gf_places if p<=10),"gl":gf_last,"sfl":sf_last,
            "bp":min(gf_places) if gf_places else None,"wp":max(gf_places) if gf_places else None,
            "bpts":max(gf_points) if gf_points else None,
            "ap":round(sum(gf_points)/len(gf_points),1) if gf_points else None,
            "apl":round(sum(gf_places)/len(gf_places),1) if gf_places else None,
            "bs":bs,"nqb":nq_best,"nqc":nq_now,"rj":reju_quals.get(nation,0),
            "ta":ta[0],"tac":ta[1],"h":hist}

    sn_list=sorted(profiles.keys(), key=lambda n:-profiles[n]["gf"])
    # Keep full history for ALL nations (no truncation)
    nation_list=[[n,profiles[n]["gf"],profiles[n]["w"]] for n in sn_list]

    with open(os.path.join(args.output, "nations.json"), "w", encoding="utf-8") as f:
        json.dump({"p":profiles,"l":nation_list}, f, ensure_ascii=False, separators=(",",":"))
    print(f"  nations.json: {len(profiles)} nations")

    # ── voting.json (only if votes file provided) ──
    if args.votes and os.path.exists(args.votes):
        import openpyxl
        wb = openpyxl.load_workbook(args.votes, read_only=True)
        ws = wb["Sheet1"]
        votes_raw = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i==0: continue
            if row[0] is None: continue
            try:
                votes_raw.append((int(row[0]),str(row[1] or "").strip(),str(row[2] or "").strip(),str(row[3] or "").strip(),int(row[4]) if row[4] else 0))
            except (ValueError, TypeError):
                continue
        wb.close()

        # Junk names to filter out (metadata from host Excel files)
        import re as _re
        _junk_patterns = [
            r"^#\s*voters?$", r"^\d+s$", r"^\d+\s*points?$", r"^sum\b", r"^count\b",
            r"^values?\s*ok$", r"^tiebreak", r"^bonus$", r"^total$", r"^points?$",
            r"^place$", r"^rank", r"^draw$", r"^result", r"^=", r"^\d+$",
            r"^check", r"^valid", r"^recap", r"^voters?$", r"^entries$",
            r"^waiting\s*list\s*jury\s*points?$",  # metadata col, not a nation
        ]
        # Build set of known short nation names from database
        _known_short = set(r.get("Nation","").lower() for r in nsc if r.get("Nation") and len(r["Nation"]) <= 2)
        def _is_junk(name):
            if not name: return True
            nl = name.lower().strip()
            if nl in _known_short: return False
            if len(name) < 2: return True
            return any(_re.search(p, nl) for p in _junk_patterns)

        # Filter out junk voter/nation names
        votes_raw = [(ed, sub, voter, nation, pts) for ed, sub, voter, nation, pts in votes_raw
                     if not _is_junk(voter) and not _is_junk(nation) and pts > 0]
        print(f"  After junk filter: {len(votes_raw)} vote records")

        # Subevent codes: 0=GF, 1=S1, 2=S2, 3=WL, 4=R1, 5=R2
        sub_code={"GF":0,"S1":1,"S2":2,"WL":3,"R1":4,"R2":5}
        # Map vote subevent to history subevent (for participant check)
        sub_v2h={"GF":"GF","S1":"SF1","S2":"SF2","WL":"GF","R1":"SF1","R2":"SF2"}
        part_set=set()
        for r in nsc: part_set.add((r["Edition"],r["Subevent"],r["Nation"]))

        # Load roster for owner mapping if available
        roster_path = args.roster if hasattr(args, 'roster') and args.roster else None
        owner_map = {}  # nation_name -> owner
        owner_aliases = {}  # owner -> [nation_names]
        if roster_path and os.path.exists(roster_path):
            import openpyxl as oxl2
            rwb = oxl2.load_workbook(roster_path, read_only=True)
            rws = rwb[rwb.sheetnames[0]]
            for ri, rrow in enumerate(rws.iter_rows(values_only=True)):
                if ri == 0: continue
                if not rrow[0]: continue
                nation_r, owner_r = str(rrow[0]).strip(), str(rrow[2] or "").strip()
                if owner_r:
                    owner_map[nation_r] = owner_r
                    if owner_r not in owner_aliases: owner_aliases[owner_r] = []
                    owner_aliases[owner_r].append(nation_r)
            rwb.close()
            print(f"  Roster loaded: {len(owner_map)} nations, {len(owner_aliases)} owners")

        # Build name normalization map (case-insensitive -> canonical)
        canon = {}
        for r in nsc:
            n = r.get("Nation","")
            if n: canon[n.lower()] = n
        # Also add roster names
        for n in owner_map:
            if n.lower() not in canon: canon[n.lower()] = n

        # WL voter name normalization — map all variants to "Waiting List"
        # Be careful NOT to match actual nations like "Waiting Iist of Shelley & Nici"
        _wl_exact = {"waiting list", "wl", "waliju", "waiting list jury"}
        def _is_wl_name(name):
            nl = name.lower().strip()
            if nl in _wl_exact: return True
            if _re.match(r"^wl\s*votes?\b", nl): return True
            if _re.match(r"^waiting\s*list$", nl): return True
            return False

        def norm(name):
            if not name: return name
            # Normalize WL voter variants
            if _is_wl_name(name): return "Waiting List"
            return canon.get(name.lower(), name)

        # Normalize names in votes
        votes_raw = [(ed, sub, norm(voter), norm(nation), pts) for ed, sub, voter, nation, pts in votes_raw]

        vote_names=sorted(set(v[2] for v in votes_raw if v[2]!="Bonus")|set(v[3] for v in votes_raw if v[2]!="Bonus"))
        vn2i={n:i for i,n in enumerate(vote_names)}
        vg=defaultdict(list)
        for ed,sub,voter,nation,pts in votes_raw:
            if voter=="Bonus": continue
            vg[(ed,sub,voter)].append((nation,pts))

        rounds=[]
        for (ed,sub,voter),vl in vg.items():
            hs=sub_v2h.get(sub,sub); isp=(ed,hs,voter) in part_set
            cat=sub_code.get(sub, 0)
            vi=vn2i.get(voter,-1)
            if vi<0: continue
            # Store RAW votes (no self-vote adjustment) + participant flag
            pairs=[]
            for nation,pts in vl:
                ni=vn2i.get(nation,-1)
                if ni>=0 and pts>0: pairs.extend([ni,pts])
            if pairs: rounds.append([ed,cat,vi,pairs,1 if isp else 0])

        love=defaultdict(lambda:[0,0])
        for ed,sub,voter,nation,pts in votes_raw:
            if voter=="Bonus" or pts<=0: continue
            vi=vn2i.get(voter,-1); ni=vn2i.get(nation,-1)
            if vi>=0 and ni>=0: love[(vi,ni)][0]+=pts; love[(vi,ni)][1]+=1
        love_list=sorted([[k[0],k[1],v[0],v[1]] for k,v in love.items() if v[1]>=3], key=lambda x:-x[2])[:300]
        ed_range=[min(v[0] for v in votes_raw),max(v[0] for v in votes_raw)]

        # Build owner merge groups for voting similarity
        owner_groups = []
        if owner_aliases:
            for owner, nations in owner_aliases.items():
                indices = [vn2i[n] for n in nations if n in vn2i]
                if len(indices) > 1:
                    owner_groups.append(indices)

        with open(os.path.join(args.output, "voting.json"), "w", encoding="utf-8") as f:
            json.dump({"r":rounds,"n":vote_names,"l":love_list,"e":ed_range,"og":owner_groups}, f, ensure_ascii=False, separators=(",",":"))
        print(f"  voting.json: {len(rounds)} rounds, {len(vote_names)} nations, {len(owner_groups)} owner groups")
    else:
        print(f"  voting.json: SKIPPED (no votes file)")

    # ── roster.json (only if roster file provided) ──
    if args.roster and os.path.exists(args.roster):
        import openpyxl as oxl3
        rwb = oxl3.load_workbook(args.roster, read_only=True)
        rws = rwb[rwb.sheetnames[0]]
        roster = []
        for ri, rrow in enumerate(rws.iter_rows(values_only=True)):
            if ri == 0: continue
            if not rrow[0]: continue
            nation_r = str(rrow[0]).strip()
            status_r = str(rrow[1] or "").strip()
            owner_r = str(rrow[2] or "").strip()
            notes_r = str(rrow[3] or "").strip() if rrow[3] else ""
            st = "current" if "urrent" in status_r else "wl" if "WL" in status_r.upper() else "defunct"
            latest = notes_r if notes_r else nation_r
            roster.append({"n": nation_r, "s": st, "o": owner_r, "ln": latest})
        rwb.close()
        # Build owner -> latest nation map
        ol = {}
        for r in roster:
            if r["o"] and r["s"] == "current": ol[r["o"]] = r["n"]
        for r in roster:
            if r["o"] and r["o"] not in ol: ol[r["o"]] = r["ln"]
        # Owner -> nations
        on = {}
        for r in roster:
            if r["o"]:
                if r["o"] not in on: on[r["o"]] = []
                on[r["o"]].append(r["n"])
        with open(os.path.join(args.output, "roster.json"), "w", encoding="utf-8") as f:
            json.dump({"r": roster, "ol": ol, "on": on}, f, ensure_ascii=False, separators=(",",":"))
        print(f"  roster.json: {len(roster)} nations ({sum(1 for r in roster if r['s']=='current')} current)")
    else:
        print(f"  roster.json: SKIPPED (no roster file)")

    total = sum(os.path.getsize(os.path.join(args.output, f)) for f in os.listdir(args.output) if f.endswith(".json"))
    print(f"\nTotal data: {total/1024:.0f} KB ({total/1024/1024:.1f} MB)")

if __name__ == "__main__":
    main()
